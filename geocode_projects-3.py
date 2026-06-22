"""
Geocoding-Skript für Mission Intelligence
==========================================

Voraussetzungen:
    pip install openpyxl requests

Ausführung:
    python3 geocode_projects.py
"""

from openpyxl import load_workbook
import time
import csv
import json
import os
import re
import requests

# SSL-Warnungen unterdrücken (Workaround für Homebrew Python / macOS)
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

INPUT_FILE = "/Users/claudiowedenig/Documents/GitHub/missionslandkarte/Backend Project Selection.xlsx"
OUTPUT_FILE = "/Users/claudiowedenig/Documents/GitHub/missionslandkarte/projects.csv"
CHECKPOINT_FILE = "/Users/claudiowedenig/Documents/GitHub/missionslandkarte/geocode_checkpoint.json"

# ── Korrekturen ────────────────────────────────────────────────────────────
CITY_CORRECTIONS = {
    'KORNEUBURG': 'Korneuburg',
    'Klagenfurt Am Wörthersee': 'Klagenfurt am Wörthersee',
    'Krems An Der Donau': 'Krems an der Donau',
    'Krems': 'Krems an der Donau',
    'Krems-Stein': 'Krems an der Donau',
    'St Polten': 'St. Pölten',
    'St. Polten': 'St. Pölten',
    'Seibersdorf An Der Leitha': 'Seibersdorf an der Leitha',
    'Kitzbuhel': 'Kitzbühel',
    'Pochlarn': 'Pöchlarn',
    'Villach-Innere Stadt': 'Villach',
}

BUNDESLAND_CORRECTIONS = {
    ('Salzburg', 'Niederösterreich'): 'Salzburg',
    ('Linz', 'Niederösterreich'): 'Oberösterreich',
    ('Wien', 'Niederösterreich'): 'Wien',
    ('Klagenfurt am Wörthersee', 'Wien'): 'Kärnten',
    ('Klagenfurt Am Wörthersee', 'Wien'): 'Kärnten',
}

# Manuelle Koordinaten für Orte die Nominatim nicht findet.
# Koordinaten über https://www.openstreetmap.org suchen.
# Format: "Stadtname|Bundesland": [lat, lng]
MANUAL_OVERRIDES = {
    "Getzersdorf bei Traismauer|Niederösterreich": [48.328, 15.696],
    "Hifstätten Raab|Steiermark": [47.067, 15.733],
    "Matreis|Tirol": [47.0, 12.533],
    "Niederranna|Kärnten": [46.835, 14.502],
}


def normalize_row(row, city_idx, bundesland_idx):
    orig_city = row[city_idx]
    orig_bl = row[bundesland_idx]
    new_city = CITY_CORRECTIONS.get(orig_city, orig_city)
    new_bl = BUNDESLAND_CORRECTIONS.get((orig_city, orig_bl), orig_bl)
    row[city_idx] = new_city
    row[bundesland_idx] = new_bl


def geocode_one(city, bundesland):
    """Geokodiert einen Standort direkt über die Nominatim HTTP API."""
    url = "https://nominatim.openstreetmap.org/search"
    params = {
        "q": f"{city}, {bundesland}, Austria",
        "format": "json",
        "limit": 1,
    }
    headers = {"User-Agent": "mission_intelligence_ffg_map_v1"}
    response = requests.get(url, params=params, headers=headers, verify=False, timeout=10)
    response.raise_for_status()
    results = response.json()
    if results:
        return round(float(results[0]["lat"]), 4), round(float(results[0]["lon"]), 4)
    return None


def geocode_all(unique_locations):
    results = {}
    failed = []

    # Checkpoint laden
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE) as f:
            results = json.load(f)
        print(f"Checkpoint geladen: {len(results)} Standorte bereits geokodiert.\n")

    for i, (city, bundesland) in enumerate(unique_locations):
        key = f"{city}|{bundesland}"

        if key in results:
            print(f"[{i+1}/{len(unique_locations)}] {city} -> bereits vorhanden, übersprungen")
            continue

        if key in MANUAL_OVERRIDES:
            results[key] = MANUAL_OVERRIDES[key]
            print(f"[{i+1}/{len(unique_locations)}] {city}, {bundesland} -> MANUAL OVERRIDE")
            with open(CHECKPOINT_FILE, 'w') as f:
                json.dump(results, f, ensure_ascii=False)
            continue

        try:
            coords = geocode_one(city, bundesland)
            if coords:
                results[key] = list(coords)
                print(f"[{i+1}/{len(unique_locations)}] {city}, {bundesland} -> OK ({coords[0]}, {coords[1]})")
            else:
                failed.append((city, bundesland))
                print(f"[{i+1}/{len(unique_locations)}] {city}, {bundesland} -> NOT FOUND")
        except Exception as e:
            failed.append((city, bundesland))
            print(f"[{i+1}/{len(unique_locations)}] {city}, {bundesland} -> ERROR: {e}")

        with open(CHECKPOINT_FILE, 'w') as f:
            json.dump(results, f, ensure_ascii=False)
        time.sleep(1.1)

    return results, failed


def slugify(text):
    text = text.lower().strip()
    for k, v in {'ä': 'ae', 'ö': 'oe', 'ü': 'ue', 'ß': 'ss'}.items():
        text = text.replace(k, v)
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')


def main():
    print(f"Lade {INPUT_FILE} ...")
    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb['Tabelle1']
    header = [cell.value for cell in ws[1]]

    type_idx       = header.index('type')
    city_idx       = header.index('city')
    bundesland_idx = header.index('bundesland')
    mission_idx    = header.index('mission')
    id_idx         = header.index('id')
    name_idx       = header.index('name')
    org_idx        = header.index('organisation')
    keywords_idx   = header.index('keywords')
    link_idx       = header.index('link')
    foerderung_idx = header.index('foerderung_eur')
    thumb_idx      = header.index('thumbnail_path')
    videolink_idx  = header.index('Videolink')
    videotype_idx  = header.index('video_type')

    point_rows = []
    pin_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[type_idx] == 'point':
            point_rows.append(list(row))
        elif row[type_idx] == 'pin':
            pin_rows.append(list(row))

    print(f"Gefunden: {len(point_rows)} Punkte, {len(pin_rows)} Pins")

    for row in point_rows + pin_rows:
        normalize_row(row, city_idx, bundesland_idx)

    unique_locations = sorted(set(
        (row[city_idx], row[bundesland_idx]) for row in point_rows + pin_rows
    ))
    print(f"\n{len(unique_locations)} eindeutige Standorte werden geokodiert ...\n")

    coords, failed = geocode_all(unique_locations)

    if failed:
        print(f"\n⚠️  {len(failed)} Standorte konnten NICHT geokodiert werden:")
        for city, bl in failed:
            print(f"   - {city}, {bl}")
        print("\nTrage diese in MANUAL_OVERRIDES oben ein und starte das Skript erneut.")
        print("Koordinaten suchen: https://www.openstreetmap.org")

    # CSV schreiben
    csv_header = [
        'id', 'type', 'mission', 'name', 'organisation', 'city', 'bundesland',
        'lat', 'lng', 'link', 'keywords',
        'foerderung_eur', 'thumbnail_path', 'video_type', 'video_id', 'video_pin_order'
    ]

    rows_written = 0
    rows_skipped = 0
    photo_counter = {}

    with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(csv_header)

        all_rows = [(r, 'point') for r in point_rows] + [(r, 'pin') for r in pin_rows]

        for row, row_type in all_rows:
            city = row[city_idx]
            bundesland = row[bundesland_idx]
            key = f"{city}|{bundesland}"

            if key not in coords:
                print(f"⚠️  Übersprungen: {row[name_idx] or row[id_idx]} ({city}, {bundesland})")
                rows_skipped += 1
                continue

            lat, lng = coords[key]
            lat_str = str(lat).replace('.', ',')
            lng_str = str(lng).replace('.', ',')

            entry_id     = row[id_idx] or slugify(row[name_idx] or "unknown")
            mission      = row[mission_idx]
            name         = row[name_idx] or ""
            organisation = (row[org_idx] or "").replace(',', ' |')
            keywords_raw = row[keywords_idx] or ""
            keywords     = " | ".join(k.strip() for k in keywords_raw.replace(';', ',').split(',') if k.strip())
            link         = row[link_idx] or ""

            foerderung = row[foerderung_idx]
            if foerderung in (None, '', 'N/A', 'n/a'):
                foerderung_str = ""
            else:
                try:
                    foerderung_str = str(int(float(str(foerderung).replace('.', '').replace(',', '.'))))
                except (ValueError, TypeError):
                    foerderung_str = ""

            if row_type == 'pin':
                thumbnail  = row[thumb_idx] or ""
                video_type = row[videotype_idx] or ""
                videolink  = row[videolink_idx] or ""
                video_id   = ""

                if video_type == 'youtube' and videolink:
                    match = re.search(r'(?:v=|/)([0-9A-Za-z_-]{11})(?:[&?/]|$)', videolink)
                    video_id = match.group(1) if match else ""
                    video_pin_order = 99
                else:
                    count = photo_counter.get(mission, 0) + 1
                    photo_counter[mission] = count
                    video_pin_order = count
            else:
                thumbnail = video_type = video_id = ""
                video_pin_order = ""

            writer.writerow([
                entry_id, row_type, mission, name, organisation, city, bundesland,
                lat_str, lng_str, link, keywords,
                foerderung_str, thumbnail, video_type, video_id, video_pin_order
            ])
            rows_written += 1

    print(f"\n✅ Fertig: {rows_written} Zeilen in {OUTPUT_FILE} geschrieben, {rows_skipped} übersprungen.")
    print(f"   Stichprobe der Koordinaten prüfen, besonders bei kleinen Orten.")
    if os.path.exists(CHECKPOINT_FILE):
        print(f"   Checkpoint-Datei kann gelöscht werden: {CHECKPOINT_FILE}")


if __name__ == "__main__":
    main()
